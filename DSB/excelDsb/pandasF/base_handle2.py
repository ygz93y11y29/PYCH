# -*- coding:utf-8 -*-

import pandas as pd

excelFile = r'cases.xlsx'
sheet_name = 'Sheet1'
truncate_sheet = False

new_infor = [
                ['#10007', 'DELL_XXXX', '2018/1/7', 'Processing', '620,000', '6.0%', 'Kate'],
                ['#10008', 'ALI_XXXX', '2018/1/8', 'Processing', '100,000', '6.0%', 'Bob'],
                ['#10009', 'Apple_XXXX', '2018/1/9', 'Pending', '80,000', '9.0%', 'Ken']
            ]


df = pd.DataFrame(new_infor, index=None)    #******************参数+++


from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 创建一个engine='openpyxl'的 ExcelWriter 对象 writer
writer = pd.ExcelWriter(excelFile, engine='openpyxl')



try:
    # 加载指定的excel文件
    writer.book = load_workbook(excelFile)

    # 得到指定sheet的最后一行数据，因为是在原excel里面添加内容
    # 所以添加的信息应该从当前sheet最后一行的后面开始
    if  sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row
        print("startrow:    ",startrow)
    else:
        startrow = 0

    # 是否需要重新创建一下该sheet
    if  truncate_sheet and sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
        writer.book.create_sheet(sheet_name, idx)

    #copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    print(writer.sheets)


except FileNotFoundError:
    # file does not exist yet, we will create it
    pass


df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False)

writer.save()








